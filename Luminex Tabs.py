import csv
import math
import tkinter as tk
from tkinter import messagebox
from tkinter.filedialog import askopenfilename, asksaveasfilename
from tkinter import *
from sys import exit
from os.path import basename, splitext
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, colors
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter

##################
# Global variables
root = tk.Tk()
##################


def main():
    file, filename = get_file()
    wb = Workbook()
    wb = conv_to_xlsx(file, wb)
    wb1 = split_tabs(wb)
    make_plate_layout(wb1)
    make_cv_table(wb1)
    choose_tabs(wb1)
    save_file(file, filename, wb1)


# Gets CSV input file from User
def get_file():
    global root
    root.withdraw()

    success = False
    while not success:
        try:
            filename = askopenfilename(title='Choose your data files',
                                       multiple=False, filetypes=(('CSV Files', '*.csv'), ('All Files', '*.*')))
            if not filename:
                exit()
            elif not filename.endswith('csv'):
                success = False
                messagebox.showerror(message="Invalid Filetype.",
                                     title="Failure")
            else:
                success = True
        except csv.Error as error:
            messagebox.showerror(message="Invalid Filetype.",
                                 title="Failure")

    file = open(filename)
    return file, filename


# Saves the new, more organized workbook
def save_file(file, filename, wb):
    new_name = splitext(basename(filename))[0]
    options = {}
    options['defaultextension'] = ".xlsx"
    options['filetypes'] = (('xlsx files', '*.xlsx'), ('all files', '*.*'))
    # options['initialdir'] = ""
    options['initialfile'] = new_name
    options['title'] = "Save as..."

    dest_filename = asksaveasfilename(**options)

    try:
        wb.save(filename=dest_filename)
    except PermissionError as e:
        messagebox.showerror(message="The file you are trying to overwrite is open. Close it and try again",
                             title="Failure")
        exit()

    if not dest_filename:
        exit()


def conv_to_xlsx(file, wb):
    # Converts file from type CSV to type XLSX to use Openpxyl module
    csv.register_dialect('comma', delimiter=',')
    reader = csv.reader(file, dialect='comma')
    ws1 = wb.worksheets[0]
    ws1.title = 'Raw data'
    for row_index, row in enumerate(reader):
        for column_index, cell in enumerate(row):
            column_letter = get_column_letter((column_index + 1))
            ws1['%s%s' % (column_letter, (row_index + 1))].value = cell
    return wb


# Divides the huge chunk of data into easier to read, separate sheets
def split_tabs(wb):
    ws = wb.worksheets[0]
    wb2 = Workbook()
    max_col = ws.max_column
    max_row = ws.max_row

    # Creates raw data sheet
    for i in range(1, max_row + 1):
        for j in range(1, max_col + 1):
            wb2.worksheets[0].cell(row=i, column=j).value = ws.cell(row=i, column=j).value
    wb2.worksheets[0].title = 'Raw Data'

    # Creates a sheet for every DataType
    counter = 1
    arr = []
    for i in range(1, max_row + 1):
        if ws.cell(row=i, column=1).internal_value == 'DataType:':
            wb2.create_sheet()
            data_type = str(ws.cell(row=i, column=2).internal_value)
            data_type = data_type.replace('/', ' & ')
            arr.append(i)
            wb2.worksheets[counter].title = data_type
            counter += 1

    # Fills in all but one DataType sheet with corresponding data
    red_font = Font(color=colors.RED)

    for i in range(0, len(arr) - 1):
        row_num = 0
        sheet = wb2.worksheets[i+1]
        for x in range(arr[i], arr[i + 1]):
            row_num += 1
            for y in range(1, max_col + 1):
                cell = sheet.cell(row=row_num, column=y)
                cell.value = ws.cell(row=x, column=y).value
                if cell.value == 'NaN':
                    cell.font = red_font

    # Fills in the last Data type sheet with corresponding data (may be a better way to do this)
    row_num = 0
    for x in range(arr[len(arr) - 1], max_row + 1):
        row_num += 1
        sheet = wb2.worksheets[len(arr)]
        for y in range(1, max_col):
            cell = sheet.cell(row=row_num, column=y)
            cell.value = ws.cell(row=x, column=y).value
            if cell.value == 'NaN':
                cell.font = red_font

    # Resize columns of all worksheets except Raw Data worksheet
    for sheet in wb2.worksheets:
        if sheet is not wb2.worksheets[0]:
            for column_cells in sheet.columns:
                def as_text(value):
                    if value is None:
                        return ""
                    return str(value)
                length = max(len(as_text(cell.value)) for cell in column_cells) + 3
                sheet.column_dimensions[column_cells[0].column].width = length

    return wb2


# Creates Plate Layout sheet
def make_plate_layout(wb):

    ws = wb.worksheets[1]
    # Finds starting number
    s = ws.cell(row=3, column=1).value
    num = int(s[-2:-1])

    arr = []
    for i in range(3, ws.max_row):
        arr.append(ws.cell(row = i, column = 2).value)

    max_col = int(len(arr)/8) + 2

    plate_layout = wb.create_sheet()
    plate_layout.title = 'Plate Layout'

    # Style
    font = Font(name='Verdana', size=12)
    bold_font = Font(name='Verdana', size=14, bold=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    for i in range(1, 10):
        plate_layout.row_dimensions[i].height = 20
        for j in range(1, max_col):
            cell = plate_layout.cell(row=i, column=j)
            cell.font = font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
    for i in range(2, 10):
        cell = plate_layout.cell(row=i, column=1)
        cell.font = bold_font
        cell.value = chr(i + 63)
    for j in range(2, max_col):
        cell = plate_layout.cell(row=1, column=j)
        cell.font = bold_font
        cell.value = num
        num += 1
        plate_layout.column_dimensions[chr(j + 64)].width = 20

    # Fills in Plate Layout sheet
    counter = 0
    for j in range(2, max_col):
        for i in range(2, 10):
            cell = plate_layout.cell(row=i, column=j)
            cell.value = arr[counter]
            cell.alignment = Alignment(horizontal='center')
            counter += 1


# Creates %CV Table Sheet
def make_cv_table(wb):
    cv_table = wb.create_sheet()
    cv_table.title = '%CV Table'
    ws = wb.worksheets[1]

    # Finds max column
    max_col = 0
    for j in range(1, ws.max_column+1):
        if ws.cell(row=2, column=j).value == 'Total Events':
            max_col = j

    # Column titles, font, and alignment
    bold_font = Font(bold=True)
    for j in range(1, max_col - 1):
        cell = cv_table.cell(row=1, column=j)
        cell.value = ws.cell(row=2, column=j+1).value
        cell.font = bold_font
        cell.alignment = Alignment(horizontal='center')

    # Column width dimensions
    for column_cells in cv_table.columns:
        cv_table.column_dimensions[column_cells[0].column].width = 20

    # Names array
    names = []
    for i in range(3, ws.max_row + 1):
        name = ws.cell(row=i, column=2).value
        if name in names:
            pass
        else:
            names.append(ws.cell(row=i, column=2).value)

    # Fills in names
    for i in range(2, len(names) + 1):
        cv_table.cell(row=i, column=1).value = names[i-2]
        cv_table.cell(row=i, column=1).font = bold_font

    # Calculates Std Dev and Mean (3 Cases: immediate duplicates, 7-gap duplicates, and no duplicates)
    std_dev = []
    mean = []
    leave = False

    # Immediate duplicates
    if ws.cell(row=3, column=2).value == ws.cell(row=4, column=2).value:
        for j in range(3, max_col):
            for i in range(3, ws.max_row, 2):
                if ws.cell(row=i, column=2).value == ws.cell(row=i+1, column=2).value:
                    p1 = ws.cell(row=i, column=j).value
                    if p1 is None:
                        break
                    p2 = ws.cell(row=i + 1, column=j).value
                    p1 = float(p1)
                    p2 = float(p2)
                    avg = (p1 + p2) / 2.0
                    mean.append(avg)
                    p1 = abs(p1 - avg)
                    p2 = abs(p2 - avg)
                    p1 *= p1
                    p2 *= p2
                    p3 = math.sqrt(p1 + p2)
                    std_dev.append(p3)
                # For some sample names with no duplicates
                else:
                    std_dev.append(999)
                    mean.append(1)

        # Calculates and fills in %CV array
        cv = []
        for i in range(0, len(std_dev)):
            x = std_dev[i] / mean[i]
            x *= 100
            x = round(x, 2)
            cv.append(x)

        # Fills in %CV into the sheet
        yellow_fill = PatternFill('solid', fgColor=colors.YELLOW)
        counter = 0
        for j in range(2, max_col + 1):
            for i in range(2, cv_table.max_row + 1):
                if counter >= len(cv):
                    pass
                else:
                    cell = cv_table.cell(row=i, column=j)
                    cell.value = cv[counter]
                    if cell.value == 0:
                        cell.fill = yellow_fill
                    counter += 1

    # 7-gap duplicates
    elif ws.cell(row=3, column=2).value == ws.cell(row=11, column=2).value:
        for j in range(3, max_col):
            visited = []
            for x in range(3, ws.max_row):
                visited.append(0)
            for i in range(3, ws.max_row):
                p1 = ws.cell(row=i, column=j).value
                if p1 is None:
                    break
                if visited[i-3] == 0:
                    if ws.cell(row=i, column=2).value == ws.cell(row=i+8, column=2).value:
                        p2 = ws.cell(row=i+8, column=j).value
                        visited[i-3] = 1
                        visited[i+5] = 1
                        p1 = float(p1)
                        p2 = float(p2)
                        avg = (p1 + p2) / 2.0
                        mean.append(avg)
                        p1 = abs(p1 - avg)
                        p2 = abs(p2 - avg)
                        p1 *= p1
                        p2 *= p2
                        p3 = math.sqrt(p1 + p2)
                        std_dev.append(p3)
                    # For some sample names with no duplicates
                    else:
                        visited[i-3] = 1
                        std_dev.append(999)
                        mean.append(1)

        # Calculates and fills in %CV array
        cv = []
        for i in range(0, len(std_dev)):
            x = std_dev[i] / mean[i]
            x *= 100
            x = round(x, 2)
            cv.append(x)

        # Fills in %CV into the sheet
        yellow_fill = PatternFill('solid', fgColor=colors.YELLOW)
        red_fill = PatternFill('solid', fgColor=colors.RED)
        counter = 0
        for j in range(2, max_col + 1):
            for i in range(2, cv_table.max_row + 1):
                if counter >= len(cv):
                    pass
                else:
                    cell = cv_table.cell(row=i, column=j)
                    cell.value = cv[counter]
                    if cell.value == 0:
                        cell.fill = yellow_fill
                    # No real %CV value
                    if cell.value == 999:
                        cell.fill = red_fill
                    counter += 1

    # Border
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    for j in range(1, max_col - 1):
        for i in range(1, cv_table.max_row + 1):
            cv_table.cell(row=i, column=j).border = thin_border


# Checkbox interface to allow user to choose which sheets they want
def choose_tabs(wb):
    names = []
    for sheet in wb.worksheets:
        names.append(sheet.title)
    master = Toplevel()

    def var_states():
        for x in range(0, len(checked)):
            if checked[x].get() == 0:
                del wb[names[x]]
        master.quit()

    Label(master, text="Select desired sheets:", anchor=CENTER).grid(row=0, sticky=W)
    checked = []
    for i in range(0, len(names)):
        var = IntVar()
        # Brendan's specified pre-selected sheets
        if (names[i] == 'Raw Data' or names[i] == 'Median' or names[i] == 'Net MFI' or names[i] == 'Count' or
                names[i] == 'Result' or names[i] == 'Avg Net MFI' or names[i] == 'Avg Result' or
                names[i] == '% Recovery' or names[i] == 'Standard Expected Concentration' or
                names[i] == 'Plate Layout' or names[i] == '%CV Table'):
            var.set(1)
        Checkbutton(master, text=names[i], variable=var).grid(row=i+1, sticky=W)
        checked.append(var)
    Button(master, text='Next', command=var_states).grid(row=len(names)+1, column=1, sticky=W, padx=0.5, pady=4)
    master.mainloop()


if __name__ == '__main__':
    exit(main())
